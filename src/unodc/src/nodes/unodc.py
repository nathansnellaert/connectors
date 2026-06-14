"""UNODC connector — thematic crime/drug statistics from data.unodc.org + WDR annexes.

One download node per collect entity (12 thematic data reports). Each fetches the
current xlsx workbook for its theme and normalizes the single main data sheet into a
universal long/tidy schema (one row per disaggregated observation), saved as parquet.

URL stability: data.unodc.org embeds a dated folder (e.g. /files/2026-05/) in file
paths that rolls forward each release, so URLs go stale. We therefore resolve the
*current* download URL at fetch time via the Drupal JSON:API file listing
(/jsonapi/file/file) keyed on the stable filename — never hardcode the dated path.
WDR annex URLs (www.unodc.org) are versioned by report year and stable within a year.

Drug data is not on the portal; the two drug themes are sourced from the flagship
World Drug Report 2025 annex workbooks (seizures for trafficking, treatment for use).
"Country profile" has no standalone dataset file on the portal (it is a PowerBI view),
so its node publishes the M49 country/region reference table that underpins it.

Strategy: stateless full re-pull. Each workbook is a few MB and the whole corpus
re-fetches in a couple of minutes, so there is no watermark/incremental machinery —
revisions and late corrections are picked up for free. Freshness gating is the
maintain step's job; if a fetch fn runs, it fetches.
"""

import io

import httpx
import openpyxl
import pyarrow as pa
from tenacity import (
    retry,
    retry_if_exception,
    stop_after_attempt,
    wait_exponential,
)

from subsets_utils import NodeSpec, SqlNodeSpec, get, save_raw_parquet

# --------------------------------------------------------------------------- #
# HTTP transport
# --------------------------------------------------------------------------- #

_TRANSIENT_EXC = (
    httpx.ConnectError,
    httpx.ConnectTimeout,
    httpx.ReadTimeout,
    httpx.WriteTimeout,
    httpx.PoolTimeout,
    httpx.RemoteProtocolError,
    httpx.ProxyError,
)


def _is_transient(exc: BaseException) -> bool:
    if isinstance(exc, _TRANSIENT_EXC):
        return True
    if isinstance(exc, httpx.HTTPStatusError):
        code = exc.response.status_code
        return code == 429 or 500 <= code < 600
    return False


@retry(
    retry=retry_if_exception(_is_transient),
    stop=stop_after_attempt(6),
    wait=wait_exponential(min=4, max=120),
    reraise=True,
)
def _request(url: str, *, params: dict | None = None) -> httpx.Response:
    resp = get(url, params=params, timeout=(15.0, 300.0))
    resp.raise_for_status()
    return resp


# --------------------------------------------------------------------------- #
# Source resolution
# --------------------------------------------------------------------------- #

PORTAL = "https://data.unodc.org"
WDR_ANNEX = "https://www.unodc.org/documents/data-and-analysis/WDR_2025/Annex/"


def _resolve_portal_url(filename: str) -> str:
    """Resolve a stable portal filename to its current dated download URL via JSON:API.

    The data.unodc.org file path embeds a YYYY-MM folder that rolls each release, so we
    look the filename up in the file listing and take the latest entry whose filename
    matches exactly (the listing's stored uri may differ, e.g. sdg_dataset.xlsx ->
    .../sdg_dataset_0.xlsx). Raises if nothing matches — a stale-only catalog must fail
    loudly, not silently fetch the wrong file.
    """
    q = (
        f"{PORTAL}/jsonapi/file/file"
        f"?filter[fn][condition][path]=filename"
        f"&filter[fn][condition][operator]=CONTAINS"
        f"&filter[fn][condition][value]={filename}"
    )
    data = _request(q).json()
    cands = []
    for f in data.get("data", []):
        a = f.get("attributes", {})
        if a.get("filename") == filename:
            uri = a.get("uri") or {}
            url = uri.get("url")
            if url:
                cands.append((a.get("changed", ""), url))
    if not cands:
        raise RuntimeError(f"no JSON:API file match for filename={filename!r}")
    cands.sort(reverse=True)
    return PORTAL + cands[0][1]


# --------------------------------------------------------------------------- #
# Universal long schema — every theme normalizes into this
# --------------------------------------------------------------------------- #

UNIVERSAL_SCHEMA = pa.schema(
    [
        ("iso3", pa.string()),
        ("geo", pa.string()),
        ("region", pa.string()),
        ("subregion", pa.string()),
        ("indicator", pa.string()),
        ("series", pa.string()),
        ("dimension", pa.string()),
        ("category", pa.string()),
        ("sex", pa.string()),
        ("age", pa.string()),
        ("drug", pa.string()),
        ("year", pa.int64()),
        ("unit", pa.string()),
        ("value", pa.float64()),
        ("value_text", pa.string()),
        ("source", pa.string()),
    ]
)
_FIELDS = [f.name for f in UNIVERSAL_SCHEMA]

_NULL_TOKENS = {"", "..", "...", "…", "-", "—", "n/a", "na", "none", "null"}


def _s(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _to_float(v) -> float | None:
    s = _s(v)
    if s is None or s.lower() in _NULL_TOKENS:
        return None
    s = s.replace(",", "").replace("%", "").strip()
    try:
        return float(s)
    except ValueError:
        return None  # censored ('<5') / annotated values kept only in value_text


def _to_year(v) -> int | None:
    s = _s(v)
    if s is None:
        return None
    s = s.split(".")[0].split("-")[0].strip()
    try:
        y = int(s)
    except ValueError:
        return None
    return y if 1800 <= y <= 2100 else None


def _row(**kw) -> dict:
    """Build a universal row with every field present (nulls for absent), so the
    parquet schema is uniform and DuckDB never trips on a missing column."""
    return {f: kw.get(f) for f in _FIELDS}


# --------------------------------------------------------------------------- #
# Workbook parsing
# --------------------------------------------------------------------------- #


def _load_main_sheet(content: bytes):
    """Return the row tuples of the first sheet. read_only mode is required (some
    workbooks carry broken drawing refs that crash the full reader); reset_dimensions
    forces a full scan because read-only dimension metadata is unreliable here."""
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    ws.reset_dimensions()
    return list(ws.iter_rows(values_only=True))


def _find_header(rows: list, anchor: str, limit: int = 20) -> int:
    anchor = anchor.lower()
    for i, row in enumerate(rows[:limit]):
        for c in row:
            if c is not None and anchor in str(c).strip().lower():
                return i
    raise RuntimeError(f"header anchor {anchor!r} not found in first {limit} rows")


def _index(header: tuple) -> dict:
    return {str(h).strip().lower(): i for i, h in enumerate(header) if h is not None}


def _cell(row: tuple, idx: dict, name: str):
    i = idx.get(name)
    if i is None or i >= len(row):
        return None
    return row[i]


def _parse_cts(rows: list, value_col: str, source_default: str) -> list[dict]:
    """CTS / GLOTIP long schema: Iso3_code, Country, Region, Subregion, Indicator,
    Dimension, Category, Sex, Age, Year, Unit of measurement, VALUE|txtVALUE, Source."""
    h = _find_header(rows, "iso3_code")
    idx = _index(rows[h])
    out = []
    for row in rows[h + 1:]:
        if not any(c is not None for c in row):
            continue
        geo = _s(_cell(row, idx, "country"))
        if geo is None:
            continue
        raw_v = _cell(row, idx, value_col)
        out.append(
            _row(
                iso3=_s(_cell(row, idx, "iso3_code")),
                geo=geo,
                region=_s(_cell(row, idx, "region")),
                subregion=_s(_cell(row, idx, "subregion")),
                indicator=_s(_cell(row, idx, "indicator")),
                dimension=_s(_cell(row, idx, "dimension")),
                category=_s(_cell(row, idx, "category")),
                sex=_s(_cell(row, idx, "sex")),
                age=_s(_cell(row, idx, "age")),
                year=_to_year(_cell(row, idx, "year")),
                unit=_s(_cell(row, idx, "unit of measurement")),
                value=_to_float(raw_v),
                value_text=_s(raw_v),
                source=_s(_cell(row, idx, "source")) or source_default,
            )
        )
    return out


def _parse_sdg(rows: list) -> list[dict]:
    h = _find_header(rows, "goal")
    idx = _index(rows[h])
    out = []
    for row in rows[h + 1:]:
        if not any(c is not None for c in row):
            continue
        geo = _s(_cell(row, idx, "geo"))
        if geo is None:
            continue
        raw_v = _cell(row, idx, "value")
        out.append(
            _row(
                iso3=_s(_cell(row, idx, "iso3_code")),
                geo=geo,
                region=_s(_cell(row, idx, "region")),
                subregion=_s(_cell(row, idx, "subregion")),
                indicator=_s(_cell(row, idx, "indicator")),
                series=_s(_cell(row, idx, "series")),
                dimension=_s(_cell(row, idx, "target")),
                category=_s(_cell(row, idx, "crime")),
                sex=_s(_cell(row, idx, "sex")),
                age=_s(_cell(row, idx, "age")),
                drug=_s(_cell(row, idx, "drug")),
                year=_to_year(_cell(row, idx, "year")),
                value=_to_float(raw_v),
                value_text=_s(raw_v),
                source=_s(_cell(row, idx, "source")) or "UNODC SDG",
            )
        )
    return out


def _parse_wildlife(rows: list) -> list[dict]:
    h = _find_header(rows, "geo")
    idx = _index(rows[h])
    out = []
    for row in rows[h + 1:]:
        if not any(c is not None for c in row):
            continue
        geo = _s(_cell(row, idx, "geo"))
        if geo is None:
            continue
        raw_v = _cell(row, idx, "value")
        out.append(
            _row(
                geo=geo,
                indicator=_s(_cell(row, idx, "indicator")),
                series=_s(_cell(row, idx, "series")),
                category=_s(_cell(row, idx, "taxonomic group")),
                year=_to_year(_cell(row, idx, "year")),
                unit=_s(_cell(row, idx, "unit_measure")),
                value=_to_float(raw_v),
                value_text=_s(raw_v),
                source=_s(_cell(row, idx, "source")) or "UNODC Wildlife",
            )
        )
    return out


def _parse_seizures(rows: list) -> list[dict]:
    """WDR 7.1 drug seizures: Region, SubRegion, Country, DrugGroup, DrugSubGroup,
    DrugName, Reference year, Kilograms, msCode."""
    h = _find_header(rows, "kilograms")
    idx = _index(rows[h])
    out = []
    for row in rows[h + 1:]:
        if not any(c is not None for c in row):
            continue
        geo = _s(_cell(row, idx, "country"))
        if geo is None:
            continue
        raw_v = _cell(row, idx, "kilograms")
        out.append(
            _row(
                iso3=_s(_cell(row, idx, "mscode")),
                geo=geo,
                region=_s(_cell(row, idx, "region")),
                subregion=_s(_cell(row, idx, "subregion")),
                indicator="Drug seizures",
                drug=_s(_cell(row, idx, "druggroup")),
                series=_s(_cell(row, idx, "drugsubgroup")),
                category=_s(_cell(row, idx, "drugname")),
                year=_to_year(_cell(row, idx, "reference year")),
                unit="Kilograms",
                value=_to_float(raw_v),
                value_text=_s(raw_v),
                source="WDR 2025 Annex 7.1",
            )
        )
    return out


def _parse_treatment(rows: list) -> list[dict]:
    """WDR 5.1 treatment by primary drug: Region, SubRegion, Country, campaign,
    questionTitle, DrugGroup, Drug, Sex, Reference year, value, ..."""
    h = _find_header(rows, "questiontitle")
    idx = _index(rows[h])
    out = []
    for row in rows[h + 1:]:
        if not any(c is not None for c in row):
            continue
        geo = _s(_cell(row, idx, "country"))
        if geo is None:
            continue
        raw_v = _cell(row, idx, "value")
        out.append(
            _row(
                geo=geo,
                region=_s(_cell(row, idx, "region")),
                subregion=_s(_cell(row, idx, "subregion")),
                indicator="Persons in drug treatment",
                dimension=_s(_cell(row, idx, "questiontitle")),
                series=_s(_cell(row, idx, "campaign")),
                drug=_s(_cell(row, idx, "druggroup")),
                category=_s(_cell(row, idx, "drug")),
                sex=_s(_cell(row, idx, "sex")),
                year=_to_year(_cell(row, idx, "reference year")),
                value=_to_float(raw_v),
                value_text=_s(raw_v),
                source="WDR 2025 Annex 5.1",
            )
        )
    return out


def _parse_regions(rows: list) -> list[dict]:
    """M49 country/region reference: ISO-alpha3 code, Country or Area, Region,
    Subregion, Intermediate Region. Underpins the 'Country profile' theme."""
    h = _find_header(rows, "iso-alpha3")
    idx = _index(rows[h])
    out = []
    for row in rows[h + 1:]:
        if not any(c is not None for c in row):
            continue
        geo = _s(_cell(row, idx, "country or area"))
        if geo is None:
            continue
        out.append(
            _row(
                iso3=_s(_cell(row, idx, "iso-alpha3 code")),
                geo=geo,
                region=_s(_cell(row, idx, "region")),
                subregion=_s(_cell(row, idx, "subregion")),
                series=_s(_cell(row, idx, "intermediate region")),
                indicator="Country classification (M49)",
                source="UNODC Data Portal M49",
            )
        )
    return out


# --------------------------------------------------------------------------- #
# Per-entity configuration (entity UUID -> how to fetch + parse)
# --------------------------------------------------------------------------- #

# group -> parser dispatch; "filename" entities resolve via JSON:API, "annex"
# entities pull a fixed (year-versioned, stable) WDR 2025 annex workbook.
ENTITY_CONFIG = {
    # 9 CTS / portal themes
    "008a2953-84ee-439a-9e9d-3853d0fb2093": {"group": "cts", "filename": "data_iafq_firearms_trafficking.xlsx", "source": "UNODC IAFQ"},
    "048e5108-306a-4add-9e71-8aac43ba2fad": {"group": "cts", "filename": "data_cts_corruption_and_economic_crime.xlsx", "source": "UNODC CTS"},
    "33bd1297-d0ec-4957-a21b-2cf35fcd2f43": {"group": "cts", "filename": "data_cts_access_and_functioning_of_justice.xlsx", "source": "UNODC CTS"},
    "69bc9e08-588a-458f-93bf-7c73a31a9efb": {"group": "cts", "filename": "data_cts_prisons_and_prisoners.xlsx", "source": "UNODC CTS"},
    "6d22d5d1-956e-404a-9d6e-499c2ed0b59e": {"group": "cts", "filename": "data_cts_violent_and_sexual_crime.xlsx", "source": "UNODC CTS"},
    "8b133928-00d0-4aa6-8bd9-4595bf47b221": {"group": "cts", "filename": "data_cts_intentional_homicide.xlsx", "source": "UNODC CTS"},
    "a0b62c57-7ffe-4a89-8f4f-19c2f99917f0": {"group": "sdg", "filename": "sdg_dataset.xlsx"},
    "bb2664cd-7ed7-4207-8527-4e05b010177f": {"group": "tip", "filename": "data_glotip.xlsx", "source": "UNODC GLOTIP"},
    "f7a38cf5-9646-4791-a7e3-46313791548d": {"group": "wildlife", "filename": "data_wildlife_trafficking.xlsx"},
    # 2 drug themes — World Drug Report 2025 annex flagship workbooks
    "8e869da8-bbe8-45fc-ada6-fcfb72dba1ca": {"group": "seizures", "annex": "7.1_Drug_seizures_2019-2023.xlsx"},
    "d0ad4f8a-f802-4a2a-bf84-440ee72e49e6": {"group": "treatment", "annex": "5.1_Treatment_by_primary_drug_of_use.xlsx"},
    # country profile — no standalone dataset; publish the M49 reference it is built on
    "21d5221e-430c-449e-aed3-0271822050d9": {"group": "regions", "filename": "data_portal_m49_regions.xlsx"},
}

# Country profile is a reference table (no observation value); everything else is a
# valued long table. Drives the transform WHERE clause.
_REFERENCE_GROUPS = {"regions"}


def _entity_id(node_id: str) -> str:
    assert node_id.startswith("unodc-"), node_id
    return node_id[len("unodc-"):]


def _parse(group: str, cfg: dict, rows: list) -> list[dict]:
    if group == "cts":
        return _parse_cts(rows, value_col="value", source_default=cfg.get("source", "UNODC"))
    if group == "tip":
        return _parse_cts(rows, value_col="txtvalue", source_default=cfg.get("source", "UNODC"))
    if group == "sdg":
        return _parse_sdg(rows)
    if group == "wildlife":
        return _parse_wildlife(rows)
    if group == "seizures":
        return _parse_seizures(rows)
    if group == "treatment":
        return _parse_treatment(rows)
    if group == "regions":
        return _parse_regions(rows)
    raise RuntimeError(f"unknown group {group!r}")


def fetch_one(node_id: str) -> None:
    asset = node_id  # the runtime hands us the spec id; it IS the asset name
    cfg = ENTITY_CONFIG[_entity_id(node_id)]
    group = cfg["group"]

    if "filename" in cfg:
        url = _resolve_portal_url(cfg["filename"])
    else:
        url = WDR_ANNEX + cfg["annex"]

    print(f"[{asset}] group={group} fetching {url}", flush=True)
    content = _request(url).content
    rows = _load_main_sheet(content)
    records = _parse(group, cfg, rows)
    if not records:
        raise RuntimeError(f"{asset}: parsed 0 records from {url}")

    table = pa.Table.from_pylist(records, schema=UNIVERSAL_SCHEMA)
    save_raw_parquet(table, asset)
    print(f"[{asset}] wrote {table.num_rows} rows", flush=True)


# --------------------------------------------------------------------------- #
# Specs
# --------------------------------------------------------------------------- #

DOWNLOAD_SPECS = [
    NodeSpec(id=f"unodc-{eid}", fn=fetch_one, kind="download")
    for eid in ENTITY_CONFIG
]


def _transform_sql(asset: str, group: str) -> str:
    cols = (
        "iso3, geo, region, subregion, indicator, series, dimension, category, "
        "sex, age, drug, year, unit, value, value_text, source"
    )
    if group in _REFERENCE_GROUPS:
        # reference table: no observation value, keyed by country
        return f'''
            SELECT DISTINCT iso3, geo, region, subregion, series, indicator, source
            FROM "{asset}"
            WHERE geo IS NOT NULL
        '''
    return f'''
        SELECT DISTINCT {cols}
        FROM "{asset}"
        WHERE geo IS NOT NULL
          AND (value IS NOT NULL OR value_text IS NOT NULL)
    '''


TRANSFORM_SPECS = [
    SqlNodeSpec(
        id=f"{spec.id}-transform",
        deps=[spec.id],
        sql=_transform_sql(spec.id, ENTITY_CONFIG[_entity_id(spec.id)]["group"]),
    )
    for spec in DOWNLOAD_SPECS
]
