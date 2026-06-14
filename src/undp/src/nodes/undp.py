"""UNDP / HDRO connector — implement step.

Two entities, two distinct bulk artefacts (stateless full re-pull each refresh;
HDRO publishes once a year, so the maintain step gates cadence):

  * composite-indices — the "All composite indices and components time series
    (1990-latest)" CSV. One row per country/aggregate (ISO3), ~1100 columns that
    are <indicator>_<year> cells (HDI, IHDI, GDI, GII, PHDI + all components).
    Wide → we melt to a long (iso3, country, hdicode, region, indicator, year,
    value) stream in the fetch fn so the SQL transform stays a thin type pass.

  * global-mpi — the global Multidimensional Poverty Index workbook
    (gMPI_Table1and2.xlsx), Table 1 = the country cross-section snapshot. xlsx is
    not SQL-readable, so we parse it (openpyxl) and emit a tidy one-row-per-country
    parquet here in Python.

File URLs embed the annual release prefix (2025_HDR / HDR25, and a YYYY-MM path
for the MPI workbook) and are NOT stable across releases, so each fetch resolves
the current link from the HDRO documentation-and-downloads page, falling back to
the verified HDR-2025 URLs if resolution fails.
"""
import csv
import io
import re

import httpx
import openpyxl
import pyarrow as pa
from tenacity import (
    retry,
    retry_if_exception,
    stop_after_attempt,
    wait_exponential,
)

from subsets_utils import (
    NodeSpec,
    SqlNodeSpec,
    get,
    load_state,
    save_raw_parquet,
    save_state,
)

STATE_VERSION = 1

DOC_URL = "https://hdr.undp.org/data-center/documentation-and-downloads"
# Verified HDR-2025 fallbacks (used only if the doc page can't be parsed).
FALLBACK_COMPOSITE_CSV = (
    "https://hdr.undp.org/sites/default/files/2025_HDR/"
    "HDR25_Composite_indices_complete_time_series.csv"
)
FALLBACK_MPI_XLSX = (
    "https://hdr.undp.org/sites/default/files/publications/additional-files/"
    "2025-10/2025_gMPI_Table1and2.xlsx"
)

# ---------------------------------------------------------------------------
# HTTP with honest retry semantics
# ---------------------------------------------------------------------------
_TRANSIENT_EXC = (
    httpx.ConnectError,
    httpx.ConnectTimeout,
    httpx.ReadTimeout,
    httpx.WriteTimeout,
    httpx.PoolTimeout,
    httpx.RemoteProtocolError,
    httpx.ProxyError,
)


def _is_transient(exc: BaseException) -> bool:
    if isinstance(exc, _TRANSIENT_EXC):
        return True
    if isinstance(exc, httpx.HTTPStatusError):
        code = exc.response.status_code
        return code == 429 or 500 <= code < 600
    return False


@retry(
    retry=retry_if_exception(_is_transient),
    stop=stop_after_attempt(6),
    wait=wait_exponential(min=4, max=120),
    reraise=True,
)
def _request(url: str) -> httpx.Response:
    resp = get(url, timeout=(10.0, 120.0))
    resp.raise_for_status()
    return resp


# ---------------------------------------------------------------------------
# Link resolution — URLs change every annual release
# ---------------------------------------------------------------------------
def _doc_links() -> tuple[list[str], list[str]]:
    html = _request(DOC_URL).text
    csvs = re.findall(
        r"https://hdr\.undp\.org/sites/default/files/[^\s\"'<>]+?\.csv", html
    )
    xlsxs = re.findall(r"https://[^\s\"'<>]+?\.xlsx", html)
    # de-dup, preserve order
    return list(dict.fromkeys(csvs)), list(dict.fromkeys(xlsxs))


def _resolve_composite_csv_url() -> str:
    try:
        csvs, _ = _doc_links()
        for u in csvs:
            if "composite" in u.lower() and "time_series" in u.lower():
                return u
        if csvs:
            return csvs[0]
    except Exception as e:  # noqa: BLE001 — resolution is best-effort; log + fall back
        print(f"composite link resolve failed {DOC_URL}: {type(e).__name__}: {e}", flush=True)
    return FALLBACK_COMPOSITE_CSV


def _resolve_mpi_xlsx_url() -> str:
    try:
        _, xlsxs = _doc_links()
        for u in xlsxs:
            if re.search(r"gmpi", u, re.I):
                return u
    except Exception as e:  # noqa: BLE001 — resolution is best-effort; log + fall back
        print(f"mpi link resolve failed {DOC_URL}: {type(e).__name__}: {e}", flush=True)
    return FALLBACK_MPI_XLSX


def _num(x):
    """Coerce a cell to float, or None for blanks / '..' / non-numeric."""
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if s in ("", "..", "n.a.", "-", "—"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# composite-indices
# ---------------------------------------------------------------------------
_COMPOSITE_META_COLS = ("iso3", "country", "hdicode", "region")

_COMPOSITE_SCHEMA = pa.schema(
    [
        ("iso3", pa.string()),
        ("country", pa.string()),
        ("hdicode", pa.string()),
        ("region", pa.string()),
        ("indicator", pa.string()),
        ("year", pa.int32()),
        ("value", pa.float64()),
    ]
)


def fetch_composite_indices(node_id: str) -> None:
    asset = node_id
    url = _resolve_composite_csv_url()
    content = _request(url).content
    # HDRO ships this file in Windows-1252, not UTF-8 (accented country names).
    text = content.decode("cp1252")
    rows = list(csv.reader(io.StringIO(text)))
    if not rows:
        raise AssertionError(f"{asset}: empty CSV from {url}")

    header = rows[0]
    if header[:4] != list(_COMPOSITE_META_COLS):
        raise AssertionError(
            f"{asset}: unexpected header start {header[:6]!r} from {url}"
        )

    # Precompute which columns are <indicator>_<year> value cells.
    melt_cols: list[tuple[int, str, int]] = []
    for ci, name in enumerate(header):
        m = re.match(r"^(.+)_(\d{4})$", name)
        if m:
            melt_cols.append((ci, m.group(1), int(m.group(2))))
    if not melt_cols:
        raise AssertionError(f"{asset}: no indicator_year columns in header")

    iso3_a, country_a, hdicode_a, region_a = [], [], [], []
    indicator_a, year_a, value_a = [], [], []

    for row in rows[1:]:
        if not row or not row[0]:
            continue
        iso3 = row[0]
        country = row[1] or None
        hdicode = row[2] or None
        region = row[3] or None
        for ci, ind, yr in melt_cols:
            v = _num(row[ci]) if ci < len(row) else None
            if v is None:
                continue
            iso3_a.append(iso3)
            country_a.append(country)
            hdicode_a.append(hdicode)
            region_a.append(region)
            indicator_a.append(ind)
            year_a.append(yr)
            value_a.append(v)

    table = pa.table(
        {
            "iso3": iso3_a,
            "country": country_a,
            "hdicode": hdicode_a,
            "region": region_a,
            "indicator": indicator_a,
            "year": year_a,
            "value": value_a,
        },
        schema=_COMPOSITE_SCHEMA,
    )
    if table.num_rows == 0:
        raise AssertionError(f"{asset}: melted to 0 rows from {url}")
    save_raw_parquet(table, asset)

    state = load_state(asset)
    state["schema_version"] = STATE_VERSION
    state["last_run_stats"] = {
        "records": table.num_rows,
        "bytes": len(content),
        "indicators": len({i for _, i, _ in melt_cols}),
        "url": url,
    }
    save_state(asset, state)


# ---------------------------------------------------------------------------
# global-mpi (Table 1 country cross-section)
# ---------------------------------------------------------------------------
# (column index in gMPI_Table1 sheet, output field). Even columns hold values;
# the interleaved odd columns are footnote-flag markers and are ignored.
_MPI_VALUE_COLS = [
    (3, "mpi_value"),
    (5, "headcount_pct"),
    (7, "pop_poor_survey_thousands"),
    (9, "pop_poor_2023_thousands"),
    (11, "intensity_pct"),
    (13, "inequality"),
    (15, "severe_poverty_pct"),
    (17, "vulnerable_pct"),
    (19, "contrib_health_pct"),
    (21, "contrib_education_pct"),
    (23, "contrib_living_standards_pct"),
    (25, "national_poverty_pct"),
    (27, "ppp300_poverty_pct"),
]

_MPI_SCHEMA = pa.schema(
    [("country", pa.string()), ("survey", pa.string())]
    + [(name, pa.float64()) for _, name in _MPI_VALUE_COLS]
)


def fetch_global_mpi(node_id: str) -> None:
    asset = node_id
    url = _resolve_mpi_xlsx_url()
    content = _request(url).content
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    if "gMPI_Table1" not in wb.sheetnames:
        raise AssertionError(
            f"{asset}: gMPI_Table1 sheet not found; sheets={wb.sheetnames} ({url})"
        )
    ws = wb["gMPI_Table1"]

    cols: dict[str, list] = {"country": [], "survey": []}
    for _, name in _MPI_VALUE_COLS:
        cols[name] = []

    n_rows = 0
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        country = row[0]
        survey = row[1] if len(row) > 1 else None
        # A data row: country name in col 0 and a survey string carrying a year
        # in col 1. Header / footnote rows fail one of these.
        if not isinstance(country, str) or not country.strip():
            continue
        if not isinstance(survey, str) or not re.search(r"\d{4}", survey):
            continue
        cols["country"].append(country.strip())
        cols["survey"].append(survey.strip())
        for ci, name in _MPI_VALUE_COLS:
            cols[name].append(_num(row[ci]) if ci < len(row) else None)
        n_rows += 1

    table = pa.table(cols, schema=_MPI_SCHEMA)
    if table.num_rows == 0:
        raise AssertionError(f"{asset}: parsed 0 data rows from {url}")
    save_raw_parquet(table, asset)

    state = load_state(asset)
    state["schema_version"] = STATE_VERSION
    state["last_run_stats"] = {"records": n_rows, "bytes": len(content), "url": url}
    save_state(asset, state)


# ---------------------------------------------------------------------------
# Specs
# ---------------------------------------------------------------------------
DOWNLOAD_SPECS = [
    NodeSpec(id="undp-composite-indices", fn=fetch_composite_indices, kind="download"),
    NodeSpec(id="undp-global-mpi", fn=fetch_global_mpi, kind="download"),
]

TRANSFORM_SPECS = [
    SqlNodeSpec(
        id="undp-composite-indices-transform",
        deps=["undp-composite-indices"],
        sql='''
            SELECT
                iso3,
                country,
                hdicode,
                region,
                indicator,
                CAST(year  AS INTEGER) AS year,
                CAST(value AS DOUBLE)  AS value
            FROM "undp-composite-indices"
            WHERE value IS NOT NULL
        ''',
    ),
    SqlNodeSpec(
        id="undp-global-mpi-transform",
        deps=["undp-global-mpi"],
        sql='''
            SELECT
                country,
                survey,
                CAST(mpi_value                    AS DOUBLE) AS mpi_value,
                CAST(headcount_pct                AS DOUBLE) AS headcount_pct,
                CAST(pop_poor_survey_thousands    AS DOUBLE) AS pop_poor_survey_thousands,
                CAST(pop_poor_2023_thousands      AS DOUBLE) AS pop_poor_2023_thousands,
                CAST(intensity_pct                AS DOUBLE) AS intensity_pct,
                CAST(inequality                   AS DOUBLE) AS inequality,
                CAST(severe_poverty_pct           AS DOUBLE) AS severe_poverty_pct,
                CAST(vulnerable_pct               AS DOUBLE) AS vulnerable_pct,
                CAST(contrib_health_pct           AS DOUBLE) AS contrib_health_pct,
                CAST(contrib_education_pct        AS DOUBLE) AS contrib_education_pct,
                CAST(contrib_living_standards_pct AS DOUBLE) AS contrib_living_standards_pct,
                CAST(national_poverty_pct         AS DOUBLE) AS national_poverty_pct,
                CAST(ppp300_poverty_pct           AS DOUBLE) AS ppp300_poverty_pct
            FROM "undp-global-mpi"
            WHERE mpi_value IS NOT NULL
        ''',
    ),
]
